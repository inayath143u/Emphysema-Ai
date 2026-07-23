import asyncio
import time
import sys
import os
import subprocess
import requests
import httpx
import statistics
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
TARGET_URL = "http://127.0.0.1:8000/"
CONCURRENCY = 100
DURATION_SECONDS = 60
OUTPUT_EXCEL = "Load_Test_Report.xlsx"

# Global list to store statistics
# Each entry: { "timestamp": float, "status_code": int, "latency": float, "success": bool, "error": str }
raw_stats = []

async def worker(client, url, stop_event):
    """
    A single virtual user that continuously sends requests until stop_event is set.
    """
    while not stop_event.is_set():
        start_time = time.perf_counter()
        timestamp = time.time()
        try:
            # Disable keep-alive to release Flask threads immediately
            headers = {"Connection": "close"}
            response = await client.get(url, headers=headers, timeout=5.0)
            latency = (time.perf_counter() - start_time) * 1000  # Convert to ms
            raw_stats.append({
                "timestamp": timestamp,
                "status_code": response.status_code,
                "latency": latency,
                "success": response.status_code == 200,
                "error": ""
            })
        except Exception as e:
            latency = (time.perf_counter() - start_time) * 1000
            err_msg = f"{type(e).__name__}: {str(e)}" if str(e) else type(e).__name__
            raw_stats.append({
                "timestamp": timestamp,
                "status_code": 0,
                "latency": latency,
                "success": False,
                "error": err_msg
            })
        
        # Simulate human user think time (150ms to 350ms) between actions
        await asyncio.sleep(random.uniform(0.15, 0.35))

async def run_load_test():
    """
    Main async load test runner that manages the virtual users.
    """
    print(f"Spawning {CONCURRENCY} virtual users...")
    stop_event = asyncio.Event()
    
    # Use httpx.AsyncClient with connection pooling enabled
    limits = httpx.Limits(max_keepalive_connections=CONCURRENCY, max_connections=CONCURRENCY * 2)
    async with httpx.AsyncClient(limits=limits) as client:
        # Start workers
        workers = [asyncio.create_task(worker(client, TARGET_URL, stop_event)) for _ in range(CONCURRENCY)]
        
        print(f"Load test started. Running for {DURATION_SECONDS} seconds...")
        
        # Run for the specified duration
        start_time = time.time()
        last_reported_second = 0
        while time.time() - start_time < DURATION_SECONDS:
            elapsed = int(time.time() - start_time)
            if elapsed > last_reported_second and elapsed % 10 == 0:
                # Periodic progress log
                total_reqs = len(raw_stats)
                print(f"[{elapsed}s / {DURATION_SECONDS}s] Sent {total_reqs} requests so far...")
                last_reported_second = elapsed
            await asyncio.sleep(0.5)
            
        # Signal workers to stop
        stop_event.set()
        
        # Wait for all workers to finish
        await asyncio.gather(*workers, return_exceptions=True)
        print("All virtual users have stopped.")

def start_flask_server():
    """
    Starts the Flask server in a subprocess.
    """
    print("Starting Flask web server on http://127.0.0.1:8000 ...")
    # We run app.py using the current Python executable
    # To prevent port locking, we ensure debug is handled or subprocess is tracked.
    env = os.environ.copy()
    env["FLASK_DEBUG"] = "0"
    server_process = subprocess.Popen(
        [sys.executable, "app.py"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env
    )
    
    # Wait for the server to be active
    server_ready = False
    print("Waiting for server to respond...")
    for _ in range(20):
        try:
            r = requests.get(TARGET_URL, timeout=1.0)
            if r.status_code == 200:
                server_ready = True
                break
        except Exception:
            pass
        time.sleep(0.5)
        
    if not server_ready:
        print("Error: Could not connect to Flask server. Terminating process...")
        terminate_process_tree(server_process.pid)
        sys.exit(1)
        
    print("Flask server successfully started.")
    return server_process

def terminate_process_tree(pid):
    """
    Kills the process tree starting from pid using taskkill on Windows.
    """
    try:
        subprocess.run(["taskkill", "/F", "/T", "/PID", str(pid)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print(f"Terminated process tree with PID {pid}")
    except Exception as e:
        print(f"Error terminating process PID {pid}: {e}")

def create_excel_report(duration_seconds, actual_duration):
    """
    Processes the raw results and generates a styled Excel report.
    """
    print(f"Generating Excel report at '{OUTPUT_EXCEL}'...")
    wb = Workbook()
    
    # Setup styles
    font_family = "Segoe UI"
    font_title = Font(name=font_family, size=16, bold=True, color="1F4E78")
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_bold = Font(name=font_family, size=11, bold=True)
    font_normal = Font(name=font_family, size=11)
    font_small = Font(name=font_family, size=9, italic=True)
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_success = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_error = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin_side = Side(style='thin', color='BFBFBF')
    border_double_side = Side(style='double', color='000000')
    
    border_cell = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_header = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=Side(style='medium', color='1F4E78'))
    border_total = Border(top=border_thin_side, bottom=border_double_side)

    # 1. SUMMARY SHEET
    ws_summary = wb.active
    ws_summary.title = "Load Test Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = "EmphysemaAI API Load Test Report"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].alignment = align_left
    ws_summary.row_dimensions[1].height = 35
    
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = font_small
    ws_summary.row_dimensions[2].height = 18
    
    # Process basic counts
    total_requests = len(raw_stats)
    successes = sum(1 for r in raw_stats if r["success"])
    failures = total_requests - successes
    success_rate = (successes / total_requests * 100) if total_requests > 0 else 0
    avg_rps = total_requests / actual_duration if actual_duration > 0 else 0
    
    latencies = [r["latency"] for r in raw_stats]
    if latencies:
        min_latency = min(latencies)
        max_latency = max(latencies)
        avg_latency = statistics.mean(latencies)
        latencies_sorted = sorted(latencies)
        p90 = latencies_sorted[int(len(latencies_sorted) * 0.90)]
        p95 = latencies_sorted[int(len(latencies_sorted) * 0.95)]
        p99 = latencies_sorted[int(len(latencies_sorted) * 0.99)]
    else:
        min_latency = max_latency = avg_latency = p90 = p95 = p99 = 0

    # Summary table data
    summary_data = [
        ("Concurrency (Virtual Users)", CONCURRENCY, "users"),
        ("Planned Test Duration", DURATION_SECONDS, "seconds"),
        ("Actual Test Duration", round(actual_duration, 2), "seconds"),
        ("Total Requests Sent", total_requests, "requests"),
        ("Successful Requests (HTTP 200)", successes, "requests"),
        ("Failed Requests", failures, "requests"),
        ("Success Rate", success_rate / 100.0, "percentage"), # Format as percentage
        ("Average Requests/Sec (RPS)", round(avg_rps, 2), "req/sec"),
        ("Minimum Response Time", round(min_latency, 1), "ms"),
        ("Average Response Time", round(avg_latency, 1), "ms"),
        ("Maximum Response Time", round(max_latency, 1), "ms"),
        ("90th Percentile Latency", round(p90, 1), "ms"),
        ("95th Percentile Latency", round(p95, 1), "ms"),
        ("99th Percentile Latency", round(p99, 1), "ms")
    ]
    
    # Write summary table headers
    ws_summary["A4"] = "Performance Metric"
    ws_summary["B4"] = "Value"
    ws_summary["C4"] = "Unit"
    
    for col in ["A4", "B4", "C4"]:
        cell = ws_summary[col]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
    ws_summary.row_dimensions[4].height = 25
    
    row_num = 5
    for metric, val, unit in summary_data:
        ws_summary.cell(row=row_num, column=1, value=metric).font = font_normal
        val_cell = ws_summary.cell(row=row_num, column=2, value=val)
        val_cell.font = font_bold
        
        # Formatting specific values
        if unit == "percentage":
            val_cell.number_format = '0.0%'
        elif unit == "requests" or unit == "users":
            val_cell.number_format = '#,##0'
        elif isinstance(val, float):
            val_cell.number_format = '#,##0.00'
        else:
            val_cell.number_format = '#,##0'
            
        val_cell.alignment = align_right
        
        ws_summary.cell(row=row_num, column=3, value=unit).font = font_small
        ws_summary.cell(row=row_num, column=3).alignment = align_center
        
        # Borders and zebra striping
        fill_row = fill_zebra if row_num % 2 == 0 else PatternFill(fill_type=None)
        for col_idx in [1, 2, 3]:
            c = ws_summary.cell(row=row_num, column=col_idx)
            c.border = border_cell
            if fill_row.fill_type:
                c.fill = fill_row
                
        ws_summary.row_dimensions[row_num].height = 20
        row_num += 1

    # 2. RPS & PERFORMANCE OVER TIME SHEET
    ws_time = wb.create_sheet(title="Performance Over Time")
    ws_time.views.sheetView[0].showGridLines = True
    
    # Headers
    time_headers = ["Elapsed Second", "Requests Sent", "Successful Requests", "Failed Requests", "Avg Latency (ms)", "RPS"]
    for col_idx, header in enumerate(time_headers, 1):
        cell = ws_time.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
    ws_time.row_dimensions[1].height = 25
    
    # Group results by second
    if raw_stats:
        start_ts = min(r["timestamp"] for r in raw_stats)
    else:
        start_ts = time.time()
        
    by_second = {}
    for r in raw_stats:
        sec = int(r["timestamp"] - start_ts) + 1
        if sec < 1:
            sec = 1
        if sec > DURATION_SECONDS:
            sec = DURATION_SECONDS
        if sec not in by_second:
            by_second[sec] = []
        by_second[sec].append(r)
        
    # Write second-by-second data
    time_row = 2
    for sec in range(1, DURATION_SECONDS + 1):
        reqs_in_sec = by_second.get(sec, [])
        count = len(reqs_in_sec)
        sec_successes = sum(1 for r in reqs_in_sec if r["success"])
        sec_failures = count - sec_successes
        sec_avg_latency = statistics.mean([r["latency"] for r in reqs_in_sec]) if count > 0 else 0
        sec_rps = count  # requests sent in this second = RPS
        
        ws_time.cell(row=time_row, column=1, value=sec).alignment = align_center
        ws_time.cell(row=time_row, column=2, value=count).number_format = '#,##0'
        ws_time.cell(row=time_row, column=3, value=sec_successes).number_format = '#,##0'
        ws_time.cell(row=time_row, column=4, value=sec_failures).number_format = '#,##0'
        ws_time.cell(row=time_row, column=5, value=round(sec_avg_latency, 2)).number_format = '#,##0.00'
        ws_time.cell(row=time_row, column=6, value=sec_rps).number_format = '#,##0'
        
        # Apply fonts and borders
        fill_row = fill_zebra if time_row % 2 == 1 else PatternFill(fill_type=None)
        for col_idx in range(1, 7):
            c = ws_time.cell(row=time_row, column=col_idx)
            c.font = font_normal
            c.border = border_cell
            if col_idx in [2, 3, 4, 5, 6]:
                c.alignment = align_right
            if fill_row.fill_type:
                c.fill = fill_row
                
        ws_time.row_dimensions[time_row].height = 20
        time_row += 1
        
    # 3. RAW REQUESTS SHEET (Limit to first 15,000 for size)
    ws_raw = wb.create_sheet(title="Raw Request Log")
    ws_raw.views.sheetView[0].showGridLines = True
    
    raw_headers = ["Request #", "Timestamp", "HTTP Status", "Latency (ms)", "Result", "Error Details"]
    for col_idx, header in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header
    ws_raw.row_dimensions[1].height = 25
    
    limit = 15000
    write_stats = raw_stats[:limit]
    
    raw_row = 2
    for idx, r in enumerate(write_stats, 1):
        ts_str = datetime.fromtimestamp(r["timestamp"]).strftime('%H:%M:%S.%f')[:-3]
        result_str = "Success" if r["success"] else "Failure"
        
        ws_raw.cell(row=raw_row, column=1, value=idx).alignment = align_center
        ws_raw.cell(row=raw_row, column=2, value=ts_str).alignment = align_center
        ws_raw.cell(row=raw_row, column=3, value=r["status_code"]).alignment = align_center
        ws_raw.cell(row=raw_row, column=4, value=round(r["latency"], 2)).number_format = '#,##0.00'
        ws_raw.cell(row=raw_row, column=4).alignment = align_right
        
        res_cell = ws_raw.cell(row=raw_row, column=5, value=result_str)
        res_cell.alignment = align_center
        if r["success"]:
            res_cell.fill = fill_success
        else:
            res_cell.fill = fill_error
            
        ws_raw.cell(row=raw_row, column=6, value=r["error"]).font = font_small
        
        for col_idx in range(1, 7):
            c = ws_raw.cell(row=raw_row, column=col_idx)
            if col_idx != 5: # Keep custom color for result
                fill_row = fill_zebra if raw_row % 2 == 1 else PatternFill(fill_type=None)
                if fill_row.fill_type:
                    c.fill = fill_row
            c.font = font_normal
            c.border = border_cell
            
        ws_raw.row_dimensions[raw_row].height = 18
        raw_row += 1
        
    if len(raw_stats) > limit:
        ws_raw.cell(row=raw_row, column=1, value=f"... and {len(raw_stats) - limit} more requests truncated ...").font = font_small
        ws_raw.merge_cells(start_row=raw_row, start_column=1, end_row=raw_row, end_column=6)

    # Auto-adjust column widths
    for ws in [ws_summary, ws_time, ws_raw]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.row == 1 and ws.title == "Load Test Summary":
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Save Workbook
    wb.save(OUTPUT_EXCEL)
    wb.close()
    print("Report saved successfully.")

def main():
    print("==================================================")
    print("EmphysemaAI Baseline & Load Testing Script")
    print("==================================================")
    
    server_process = start_flask_server()
    
    actual_duration = 0
    start_time = time.perf_counter()
    try:
        # Run async load test
        asyncio.run(run_load_test())
        actual_duration = time.perf_counter() - start_time
        print(f"Load test execution finished in {actual_duration:.2f} seconds.")
    except KeyboardInterrupt:
        print("Load test interrupted by user.")
    except Exception as e:
        print(f"Unexpected error during load test: {e}")
    finally:
        # Terminate server
        print("Terminating Flask server...")
        terminate_process_tree(server_process.pid)
        
    # Process data and generate report
    if raw_stats:
        total_requests = len(raw_stats)
        successes = sum(1 for r in raw_stats if r["success"])
        failures = total_requests - successes
        avg_rps = total_requests / actual_duration if actual_duration > 0 else 0
        latencies = [r["latency"] for r in raw_stats]
        avg_lat = statistics.mean(latencies) if latencies else 0
        min_lat = min(latencies) if latencies else 0
        max_lat = max(latencies) if latencies else 0
        
        print("\n==================================================")
        print("Load Test Summary Results")
        print("==================================================")
        print(f"Requests per Second (RPS): {avg_rps:.2f} req/sec")
        print(f"Total Requests Sent:       {total_requests} (Success: {successes}, Failures: {failures})")
        print(f"Success Rate:              {(successes / total_requests * 100):.2f}%")
        print("Response Time (Latency):")
        print(f"  Average: {avg_lat:.2f}ms")
        print(f"  Minimum: {min_lat:.2f}ms")
        print(f"  Maximum: {max_lat:.2f}ms")
        print("==================================================")
        
        create_excel_report(DURATION_SECONDS, actual_duration)
    else:
        print("No requests were recorded. Excel report will not be generated.")

if __name__ == "__main__":
    main()
