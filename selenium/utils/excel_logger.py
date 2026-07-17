import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

class ExcelLogger:
    @staticmethod
    def log_test_result(test_name, status, error_message="-", screenshot_path="-"):
        reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        excel_file = os.path.join(reports_dir, 'test_results.xlsx')

        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Run Summary"
            # Write Header
            ws.append(["Timestamp", "Test Case Name", "Status", "Error Details", "Screenshot Link"])

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, test_name, status, error_message, screenshot_path])
        wb.save(excel_file)
        wb.close()
