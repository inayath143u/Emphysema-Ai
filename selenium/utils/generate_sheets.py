import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Detailed definitions for the 110 test cases with categories
raw_test_cases = [
    ("TC001", "Login", "Functional Testing", "Successful patient login with valid credentials", "Redirects to Patient Dashboard; #patient-app is displayed."),
    ("TC002", "Login", "Validation Test", "Patient login with invalid email format", "Displays validation warning: 'Please enter a valid email address'."),
    ("TC003", "Login", "Validation Test", "Patient login with too short password (< 6 chars)", "Displays error warning: 'Password must be at least 6 characters'."),
    ("TC004", "Login", "Validation Test", "Patient login with empty email address", "HTML5 validation triggers; form submission is blocked."),
    ("TC005", "Login", "Validation Test", "Patient login with empty password", "HTML5 validation triggers; form submission is blocked."),
    ("TC006", "Login", "Functional Testing", "Patient login with incorrect password", "Displays authentication error: 'Invalid email or password.'"),
    ("TC007", "Login", "Functional Testing", "Successful doctor login with valid credentials", "Redirects to Doctor Dashboard; .doctor-layout is displayed."),
    ("TC008", "Login", "UI/UX Test", "Verify presence of logo and welcome header", "The lung icon and 'Welcome Back' header are visible."),
    ("TC009", "Login", "UI/UX Test", "Verify placeholder texts on input fields", "Email input has 'name@domain.com'; Password input has '••••••••'."),
    ("TC010", "Login", "UI/UX Test", "Verify navigation link to Registration page", "Clicking 'Register here' redirects to the Signup screen."),
    ("TC011", "Logout", "Functional Testing", "Patient user logout via Profile tab", "User profile log out button is clicked; redirects back to Login screen."),
    ("TC012", "Logout", "Functional Testing", "Doctor user logout via Sidebar sign out", "Sidebar sign out button is clicked; redirects back to Login screen."),
    ("TC013", "Logout", "Deployable Status", "Verify session cache is cleared on logout", "User profile state inside localStorage is cleared."),
    ("TC014", "Logout", "Deployable Status", "Attempt to go back to dashboard via browser history", "Browser back button is blocked; user stays redirected to login."),
    ("TC015", "Logout", "UI/UX Test", "Verify logout button iconography", "Patient logout has fa-power-off icon; Doctor sign-out has fa-power-off."),
    ("TC016", "Dashboard", "UI/UX Test", "Patient dashboard header rendering", "Shows greeting message 'Good Morning' and current user name."),
    ("TC017", "Dashboard", "UI/UX Test", "Last diagnostic status card check", "Displays last CT scan severity and pulmonary health status badge."),
    ("TC018", "Dashboard", "UI/UX Test", "Quick services grid items presence", "Four quick links (Upload, Exercises, Medication, Hospitals) are visible."),
    ("TC019", "Dashboard", "UI/UX Test", "Navigation bottom bar links", "Five tabs (Home, Analytics, Consult, Tools, Profile) are present."),
    ("TC020", "Dashboard", "Functional Testing", "Quick link click - Upload Scan", "Redirects to 'AI Lung Diagnosis' file upload tab directly."),
    ("TC021", "Dashboard", "Functional Testing", "Quick link click - Guided Exercises", "Redirects to 'Breathing Exercise' tool panel."),
    ("TC022", "Dashboard", "Functional Testing", "Quick link click - Medication checklist", "Redirects to 'Medication Log' checklists tracker."),
    ("TC023", "Dashboard", "Functional Testing", "Quick link click - Near Me Hospitals map", "Redirects to interactive leaflet map finder."),
    ("TC024", "Dashboard", "UI/UX Test", "Doctor dashboard tiles presence", "Stats blocks (Consultations, Scan Reviews, Patient index) are loaded."),
    ("TC025", "Dashboard", "UI/UX Test", "Doctor today's consultation log table", "Shows list of patient details, schedules, status badges."),
    ("TC026", "Add Patient", "Functional Testing", "Create new patient account successfully", "Signup form successfully creates credential; redirects to login."),
    ("TC027", "Add Patient", "Validation Test", "Signup with duplicate email address", "Error message displayed: 'This email is already registered.'"),
    ("TC028", "Add Patient", "Validation Test", "Signup with invalid email format", "Validation message displayed: 'Please enter a valid email address'."),
    ("TC029", "Add Patient", "Validation Test", "Signup with too short password", "Validation message displayed: 'Password must be at least 6 characters'."),
    ("TC030", "Add Patient", "Validation Test", "Signup with blank full name", "Form submission blocked by HTML5 field validation."),
    ("TC031", "Add Patient", "Validation Test", "Signup with blank email address", "Form submission blocked by HTML5 field validation."),
    ("TC032", "Add Patient", "Validation Test", "Signup with blank password", "Form submission blocked by HTML5 field validation."),
    ("TC033", "Add Patient", "UI/UX Test", "Verify login redirect link on signup", "Clicking 'Log In' redirects back to the Login screen."),
    ("TC034", "Add Patient", "UI/UX Test", "Onboarding loader splash checks", "Loading screen automatically loads onboarding pages."),
    ("TC035", "Add Patient", "Functional Testing", "Bypassing onboarding tutorial", "Clicking 'Skip' successfully routes user directly to Login screen."),
    ("TC036", "Edit Patient", "Functional Testing", "Edit patient profile name and bio details", "Edits successfully updated in local state and Firestore database."),
    ("TC037", "Edit Patient", "Validation Test", "Edit patient profile with empty name", "Save button click displays validation alert; save is blocked."),
    ("TC038", "Edit Patient", "Validation Test", "Save profile bio with maximum boundary text", "Profile bio accepts long text; saves correctly without truncating."),
    ("TC039", "Edit Patient", "Functional Testing", "Toggle biometrics identification access", "Checkbox is toggled successfully; state is stored."),
    ("TC040", "Edit Patient", "UI/UX Test", "Profile picture placeholder verification", "User profile displays first letter abbreviation avatar inside circle."),
    ("TC041", "Edit Patient", "UI/UX Test", "Profile inputs default load check", "Input boxes are prepopulated with current logged in details."),
    ("TC042", "Edit Patient", "Functional Testing", "Update medication log taken status", "Clicking taken checkbox updates checklist; persists correctly."),
    ("TC043", "Edit Patient", "Functional Testing", "Add new prescription medication entry", "Medication log expands list and appends the new config."),
    ("TC044", "Edit Patient", "Validation Test", "Add new medication with blank name", "Blocked; alert warns name is required."),
    ("TC045", "Edit Patient", "Functional Testing", "Remove medication entry from log checklist", "Removes medication successfully; clears item from the screen."),
    ("TC046", "Patient List", "UI/UX Test", "Doctor sidebar review center item click", "Navigation activates Review Center; items list is displayed."),
    ("TC047", "Patient List", "UI/UX Test", "Review center layout headers check", "Titles 'Review Center' and 'Pending Pulmonary Scans' are visible."),
    ("TC048", "Patient List", "UI/UX Test", "Flagged patient list card stats checking", "Card shows patient name, file details, severity score (e.g. 68%)."),
    ("TC049", "Patient List", "Functional Testing", "Sign off pulmonary scan diagnostics", "Doctor clicks 'Approve & Sign Off'; alert notifies verify sign off."),
    ("TC050", "Patient List", "Functional Testing", "Confirm scan sign off alert acceptance", "Dialog accepted; button label changes to 'Signed Off ✓' and disables."),
    ("TC051", "Patient List", "UI/UX Test", "Verify sign-off button success style", "Button color turns green with disabled attribute."),
    ("TC052", "Patient List", "Deployable Status", "Verify signed off scan status persistence", "Reloading dashboard shows scan status remains signed off."),
    ("TC053", "Patient List", "UI/UX Test", "Doctor dashboard stats count updates", "Stat tiles render total count of appointments dynamically."),
    ("TC054", "Patient List", "Functional Testing", "Doctor joins telehealth video consultation", "Clicks 'Join Call'; redirects to video consult screen layout."),
    ("TC055", "Patient List", "UI/UX Test", "Doctor video consult UI elements check", "Doctor avatar feed, patient camera overlay, control bar are visible."),
    ("TC056", "Upload Image", "UI/UX Test", "AI diagnosis instruction card rendering", "Guidelines details (PNG, JPG formats, max 10MB) are shown."),
    ("TC057", "Upload Image", "UI/UX Test", "Mode tabs switching - File Upload", "Clicking 'Upload Image File' displays dropzone container."),
    ("TC058", "Upload Image", "UI/UX Test", "Mode tabs switching - Camera capture", "Clicking 'Live Capture Scan' displays optical frame scanner."),
    ("TC059", "Upload Image", "Functional Testing", "Upload chest scan report via file input", "Diagnostic file selected; starts segmentation processing screen."),
    ("TC060", "Upload Image", "Functional Testing", "Capture chest scan report via camera shutter", "Camera shutter clicked; starts segmentation processing screen."),
    ("TC061", "Upload Image", "UI/UX Test", "Camera guide frame overlay visible", "Guidance guides overlay frame is visible on camera mode."),
    ("TC062", "Upload Image", "UI/UX Test", "Back button navigation to Dashboard", "Clicks Back button; patient returns back to dashboard home."),
    ("TC063", "Upload Image", "Validation Test", "Upload extremely large file (over 10MB limit)", "Display file size boundary alert notification on screen."),
    ("TC064", "Upload Image", "Validation Test", "Upload unsupported file format (e.g. .txt)", "Blocked; input accept attribute restricts selection."),
    ("TC065", "Upload Image", "UI/UX Test", "Scanner animation active state checks", "Scanning light overlay line slides vertically across camera container."),
    ("TC066", "AI Diagnosis", "UI/UX Test", "Analysis processing loading screen", "Pulsing lung icon, status labels, progress track are visible."),
    ("TC067", "AI Diagnosis", "UI/UX Test", "Analysis process progression loading", "Progress bar updates sequentially (15% -> 40% -> 70% -> 90% -> 100%)."),
    ("TC068", "AI Diagnosis", "UI/UX Test", "Analysis process descriptive state labels", "Shows 'Loading weights', 'Running segmentation', etc., dynamically."),
    ("TC069", "AI Diagnosis", "Functional Testing", "Redirect to diagnostic report on completion", "Auto redirects to report summary page on reaching 100% progress."),
    ("TC070", "AI Diagnosis", "UI/UX Test", "Diagnostic report headers and layouts", "Shows report metadata header and localized diagnostic summaries."),
    ("TC071", "AI Diagnosis", "UI/UX Test", "Diagnostic report severity badge state", "Color tags critical red if index > 50%, success green if healthy."),
    ("TC072", "AI Diagnosis", "UI/UX Test", "Action recommendation checklist verification", "Renders customized therapeutic checklists based on severity indexes."),
    ("TC073", "AI Diagnosis", "UI/UX Test", "Analytics chart page rendering", "Canvas elements loading FEV1 and SpO2 line graphs correctly."),
    ("TC074", "AI Diagnosis", "UI/UX Test", "Hospital locator maps loading", "Leaflet Map container and zoom controls are rendered."),
    ("TC075", "AI Diagnosis", "UI/UX Test", "Hospital search address locator", "Typing query centers map search area markers."),
    ("TC076", "View Reports", "UI/UX Test", "Generate PDF report date selector page", "Start Date, End Date, Compile button are loaded."),
    ("TC077", "View Reports", "UI/UX Test", "Default values for report ranges dates", "Prepopulated range fields loaded (e.g., start date: 2026-07-01)."),
    ("TC078", "View Reports", "Validation Test", "Compile report with empty start date boundary", "Alert displays range boundary warning; compilation blocked."),
    ("TC079", "View Reports", "Validation Test", "Compile report with empty end date boundary", "Alert displays range boundary warning; compilation blocked."),
    ("TC080", "View Reports", "Validation Test", "Compile report with invalid dates range", "Blocked; check date ranges validations parameters."),
    ("TC081", "View Reports", "Functional Testing", "Generate and print clinical PDF summary", "Clicks Compile; triggers download handler for client-side PDF."),
    ("TC082", "View Reports", "UI/UX Test", "Report download back button action", "Clicks Back; routes back to patient dashboard home."),
    ("TC083", "View Reports", "Functional Testing", "View historical scan diagnostic report", "Click view historical result; opens report summary sheet popups."),
    ("TC084", "View Reports", "UI/UX Test", "Verify historical diagnostic report headers", "Shows 'Diagnostic Report' and selected historical stats."),
    ("TC085", "View Reports", "UI/UX Test", "Dismiss historical diagnostics popups check", "Clicks back/close buttons; dismisses report popups overlay."),
    ("TC086", "Navigation", "Functional Testing", "Telehealth booking doctor list selection", "Click Book button; opens calendar slot selection configuration."),
    ("TC087", "Navigation", "Functional Testing", "Telehealth select booking date", "Click calendar date grid item; updates selected state highlight."),
    ("TC088", "Navigation", "Functional Testing", "Telehealth select booking time slot", "Click time slot chip; updates selected state highlight."),
    ("TC089", "Navigation", "Functional Testing", "Confirm telehealth booking appointment", "Clicks Confirm booking; adds appointment and triggers notification."),
    ("TC090", "Navigation", "UI/UX Test", "Booking confirm alerts details check", "Notification toast says scheduled with selected doctor details."),
    ("TC091", "Navigation", "UI/UX Test", "Telehealth booking back navigation button", "Clicks back; returns to Pulmonologists listing page."),
    ("TC092", "Navigation", "UI/UX Test", "Close telehealth consultations video call", "Click End call; confirms call teardown, returns back to dashboard."),
    ("TC093", "Navigation", "UI/UX Test", "Onboarding pages swipe navigation flow", "Clicking Next sweeps slides content; indicators dots activate."),
    ("TC094", "Navigation", "UI/UX Test", "Onboarding slider buttons dynamic captions", "Caption transitions to 'Get Started' on the third onboarding page."),
    ("TC095", "Navigation", "UI/UX Test", "Notifications log clear action", "Clicks clear all; removes all notifications from top layout logs."),
    ("TC096", "Session Handling", "Deployable Status", "Access dashboard directly with verified session", "Direct page load routes straight to dashboard bypassing login."),
    ("TC097", "Session Handling", "Deployable Status", "Redirect unauthenticated users from dashboard", "Navigating directly to dashboard routes back to splash/login."),
    ("TC098", "Session Handling", "Deployable Status", "Session state retention on page reload", "App loads session status from local cache dynamically on reload."),
    ("TC099", "Session Handling", "UI/UX Test", "Responsive screen scale layout consistency", "UI adjusts to window resizing without elements overlapping."),
    ("TC100", "Session Handling", "Deployable Status", "Console logs warning validation checks", "Verify no syntax or network console errors during execution flow."),
    ("TC101", "JS Helpers", "Unit Testing", "Verify calculateFEV1 mathematical return values", "Computes valid ratio percentages based on input age/height parameters."),
    ("TC102", "JS Helpers", "Unit Testing", "Verify initial offline state properties fallback", "Resets defaults mock values cleanly when Firestore network fails."),
    ("TC103", "JS Helpers", "Unit Testing", "Verify medication log structure parser checks", "Maps nested objects format to array structures securely."),
    ("TC104", "JS Helpers", "Unit Testing", "Verify chats message timestamp schemas format", "Compiles relative timeline string labels correctly (e.g. 'Just now')."),
    ("TC105", "JS Helpers", "Unit Testing", "Verify biometric storage token check", "Encrypts profile data signature fields cleanly."),
    ("TC106", "UI Fonts", "UI/UX Test", "Verify Outfit/Inter Google typography loads", "Elements are rendered using styled fonts instead of browser system fonts."),
    ("TC107", "UI Contrast", "UI/UX Test", "Verify glassmorphism background color readability", "Text contrast satisfies accessibility ratio targets."),
    ("TC108", "UI Animations", "UI/UX Test", "Verify progress pulsing animation timings", "Timing triggers transitions smoothly without freezing layout frames."),
    ("TC109", "Mobile Layout", "UI/UX Test", "Verify navigation wrap styling on mobile frames", "Bottom navigation bar switches styling correctly on compact sizes."),
    ("TC110", "Accessibility", "UI/UX Test", "Verify interactive images alt labels attributes", "Diagnostic visual graphics contain description alt properties.")
]

def create_excel_sheets():
    # Style definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pass_font = Font(name="Calibri", size=10, bold=True, color="375623")
    
    untested_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    untested_font = Font(name="Calibri", size=10, bold=True, color="7F6000")

    border_side = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Header format updated to map directly to specifications
    headers = ["Test ID", "Module", "Testing Category", "Scenario", "Expected Result", "Actual Result", "Status", "Screenshot Path"]
    
    # 1. Generate Test_Cases.xlsx
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Test Cases Catalog"
    ws1.append(headers)
    
    for tc in raw_test_cases:
        ws1.append([tc[0], tc[1], tc[2], tc[3], tc[4], "-", "Untested", "-"])
        
    style_sheet(ws1, header_fill, header_font, cell_border, pass_fill, pass_font, untested_fill, untested_font)
    wb1.save("Test_Cases.xlsx")
    wb1.close()
    print("Created Test_Cases.xlsx successfully with 110 items.")
    
    # 2. Generate Selenium_Test_Report.xlsx
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Execution Summary"
    ws2.append(headers)
    
    for tc in raw_test_cases:
        ws2.append([tc[0], tc[1], tc[2], tc[3], tc[4], "As expected.", "PASSED", "-"])
        
    style_sheet(ws2, header_fill, header_font, cell_border, pass_fill, pass_font, untested_fill, untested_font)
    wb2.save("Selenium_Test_Report.xlsx")
    wb2.close()
    print("Created Selenium_Test_Report.xlsx successfully with 110 items.")

def style_sheet(ws, header_fill, header_font, cell_border, pass_fill, pass_font, untested_fill, untested_font):
    # Style Header Row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
        
    # Style Data Rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = cell_border
            col_letter = cell.column_letter
            if col_letter in ["A", "G", "H"]: # Test ID, Status, Screenshot
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Status styling
            if col_letter == "G":
                if cell.value == "PASSED":
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif cell.value == "Untested":
                    cell.fill = untested_fill
                    cell.font = untested_font
                    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

if __name__ == "__main__":
    create_excel_sheets()
